"""
Extrator de Currículos Lattes
Lê arquivos .txt (HTML do Lattes) e extrai: nome, formação,
projetos de pesquisa, projetos de ensino e projetos de extensão.
Salva tudo em uma planilha .xlsx local.

Uso:
    python extrair_lattes.py                    # lê todos os .txt da pasta atual
    python extrair_lattes.py pasta/com/arquivos # lê .txt de outra pasta

Dependências:
    pip install beautifulsoup4 openpyxl
"""

import os
import sys
import glob
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ---------------------------------------------------------------------------
# Funções de extração
# ---------------------------------------------------------------------------

def carregar_html(caminho):
    with open(caminho, "r", encoding="utf-8", errors="replace") as f:
        return BeautifulSoup(f.read(), "html.parser")


def extrair_nome(soup):
    tag = soup.find("h2", class_="nome")
    return tag.get_text(strip=True) if tag else ""


def extrair_formacao(soup):
    ancora = soup.find("a", {"name": "FormacaoAcademicaTitulacao"})
    if not ancora:
        return []
    wrapper = ancora.find_parent("div", class_="title-wrapper")
    if not wrapper:
        return []
    data_cell = wrapper.find("div", class_="data-cell")
    if not data_cell:
        return []
    periodos = data_cell.find_all("div", class_="layout-cell-3")
    conteudos = data_cell.find_all("div", class_="layout-cell-9")
    itens = []
    for periodo, conteudo in zip(periodos, conteudos):
        b = periodo.find("b")
        per_txt = b.get_text(strip=True) if b else ""
        desc = " ".join(conteudo.get_text(" ", strip=True).split())
        itens.append({"periodo": per_txt, "descricao": desc})
    return itens


def _extrair_projetos_por_ancora(soup, nome_ancora):
    ancora = soup.find("a", {"name": nome_ancora})
    if not ancora:
        return []
    wrapper = ancora.find_parent("div", class_="title-wrapper")
    if not wrapper:
        return []
    data_cell = wrapper.find("div", class_="data-cell")
    if not data_cell:
        return []
    periodos = data_cell.find_all("div", class_="layout-cell-3")
    conteudos = data_cell.find_all("div", class_="layout-cell-9")
    projetos = []
    for periodo, conteudo in zip(periodos, conteudos):
        b = periodo.find("b")
        per_txt = b.get_text(strip=True) if b else ""
        titulo_tag = conteudo.find("div", class_="layout-cell-pad-5")
        titulo = ""
        if titulo_tag:
            titulo_b = titulo_tag.find("b")
            titulo = titulo_b.get_text(strip=True) if titulo_b else ""
        desc = " ".join(conteudo.get_text(" ", strip=True).split())
        projetos.append({"periodo": per_txt, "titulo": titulo, "descricao": desc})
    return projetos


def extrair_projetos_pesquisa(soup):
    return _extrair_projetos_por_ancora(soup, "ProjetosPesquisa")


def extrair_projetos_extensao(soup):
    return _extrair_projetos_por_ancora(soup, "ProjetosExtensao")


def extrair_projetos_ensino(soup):
    """
    Busca projetos de ensino formais (âncora 'ProjetosEnsino').
    Retorna lista vazia se a seção não existir no currículo.
    """
    ancora = soup.find("a", {"name": "ProjetosEnsino"})
    if not ancora:
        return []
    wrapper = ancora.find_parent("div", class_="title-wrapper")
    if not wrapper:
        return []
    data_cell = wrapper.find("div", class_="data-cell")
    if not data_cell:
        return []
    periodos = data_cell.find_all("div", class_="layout-cell-3")
    conteudos = data_cell.find_all("div", class_="layout-cell-9")
    itens = []
    for periodo, conteudo in zip(periodos, conteudos):
        b = periodo.find("b")
        per_txt = b.get_text(strip=True) if b else ""
        titulo_b = conteudo.find("b")
        titulo = titulo_b.get_text(strip=True) if titulo_b else ""
        desc = " ".join(conteudo.get_text(" ", strip=True).split())
        itens.append({"periodo": per_txt, "titulo": titulo, "descricao": desc})
    return itens


# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------

AZUL_HEADER     = "2F5496"
AZUL_ESCURO     = "1F3864"
CINZA_SUBHEADER = "D9E1F2"

def estilo_header(cell, bg=AZUL_HEADER):
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def estilo_subheader(cell):
    cell.font = Font(name="Arial", bold=True, color=AZUL_ESCURO, size=10)
    cell.fill = PatternFill("solid", start_color=CINZA_SUBHEADER)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def estilo_dado(cell, wrap=True):
    cell.font = Font(name="Arial", size=10)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)

def aplicar_borda(cell):
    lado = Side(style="thin", color="BFBFBF")
    cell.border = Border(left=lado, right=lado, top=lado, bottom=lado)


# ---------------------------------------------------------------------------
# Abas da planilha
# ---------------------------------------------------------------------------

def criar_planilha_resumo(wb, todos_dados):
    ws = wb.create_sheet("Resumo", 0)
    headers = [
        "Pesquisador", "Qtd. Formações", "Qtd. Proj. Pesquisa",
        "Qtd. Proj. Ensino", "Qtd. Proj. Extensão", "Arquivo",
    ]
    ws.append(headers)
    for cell in ws[1]:
        estilo_header(cell)
        aplicar_borda(cell)
    for d in todos_dados:
        ws.append([
            d["nome"], len(d["formacao"]), len(d["pesquisa"]),
            len(d["ensino"]), len(d["extensao"]), d["arquivo"],
        ])
        for cell in ws[ws.max_row]:
            estilo_dado(cell, wrap=False)
            aplicar_borda(cell)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 40
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def criar_aba_formacao(wb, todos_dados):
    ws = wb.create_sheet("Formação")
    ws.append(["Pesquisador", "Período", "Descrição"])
    for cell in ws[1]:
        estilo_header(cell)
        aplicar_borda(cell)
    for d in todos_dados:
        for item in d["formacao"]:
            ws.append([d["nome"], item["periodo"], item["descricao"]])
            for cell in ws[ws.max_row]:
                estilo_dado(cell)
                aplicar_borda(cell)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 90
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def criar_aba_projetos(wb, nome_aba, todos_dados, chave):
    ws = wb.create_sheet(nome_aba)
    ws.append(["Pesquisador", "Período", "Título", "Descrição Completa"])
    for cell in ws[1]:
        estilo_header(cell)
        aplicar_borda(cell)
    for d in todos_dados:
        for item in d[chave]:
            ws.append([
                d["nome"],
                item.get("periodo", ""),
                item.get("titulo", ""),
                item.get("descricao", ""),
            ])
            for cell in ws[ws.max_row]:
                estilo_dado(cell)
                aplicar_borda(cell)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 90
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def criar_aba_pesquisador(wb, dados):
    """Aba individual com todas as informações do pesquisador."""
    ws = wb.create_sheet(dados["nome"][:28].strip())
    linha = 1

    ws.cell(linha, 1, "CURRÍCULO LATTES")
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
    estilo_header(ws.cell(linha, 1))
    linha += 1

    for rotulo, valor in [("Pesquisador:", dados["nome"]), ("Arquivo:", dados["arquivo"])]:
        ws.cell(linha, 1, rotulo)
        ws.cell(linha, 2, valor)
        estilo_subheader(ws.cell(linha, 1))
        estilo_dado(ws.cell(linha, 2), wrap=False)
        ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=4)
        linha += 1
    linha += 1

    secoes = [
        ("FORMAÇÃO ACADÊMICA",   dados["formacao"], ["Período", "Descrição"],           ["periodo", "descricao"]),
        ("PROJETOS DE PESQUISA", dados["pesquisa"],  ["Período", "Título", "Descrição"],  ["periodo", "titulo", "descricao"]),
        ("PROJETOS DE ENSINO",   dados["ensino"],    ["Período", "Título", "Descrição"],  ["periodo", "titulo", "descricao"]),
        ("PROJETOS DE EXTENSÃO", dados["extensao"],  ["Período", "Título", "Descrição"],  ["periodo", "titulo", "descricao"]),
    ]

    for titulo_secao, itens, cols, chaves in secoes:
        ws.cell(linha, 1, titulo_secao)
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=len(cols))
        estilo_header(ws.cell(linha, 1), bg=AZUL_ESCURO)
        linha += 1
        for i, col in enumerate(cols, 1):
            c = ws.cell(linha, i, col)
            estilo_subheader(c)
            aplicar_borda(c)
        linha += 1
        if itens:
            for item in itens:
                for i, chave in enumerate(chaves, 1):
                    c = ws.cell(linha, i, item.get(chave, ""))
                    estilo_dado(c)
                    aplicar_borda(c)
                linha += 1
        else:
            ws.cell(linha, 1, "Nenhuma informação encontrada.")
            ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=len(cols))
            linha += 1
        linha += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 90
    ws.column_dimensions["D"].width = 12


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    pasta = sys.argv[1] if len(sys.argv) > 1 else "."
    arquivos = glob.glob(os.path.join(pasta, "*.txt"))

    if not arquivos:
        print(f"Nenhum arquivo .txt encontrado em: {os.path.abspath(pasta)}")
        sys.exit(1)

    print(f"Encontrados {len(arquivos)} arquivo(s) .txt\n")

    todos_dados = []
    for caminho in sorted(arquivos):
        nome_arquivo = os.path.basename(caminho)
        print(f"  Processando: {nome_arquivo} ...", end=" ")
        try:
            soup = carregar_html(caminho)
            dados = {
                "arquivo":  nome_arquivo,
                "nome":     extrair_nome(soup) or nome_arquivo,
                "formacao": extrair_formacao(soup),
                "pesquisa": extrair_projetos_pesquisa(soup),
                "ensino":   extrair_projetos_ensino(soup),
                "extensao": extrair_projetos_extensao(soup),
            }
            todos_dados.append(dados)
            print("OK")
        except Exception as e:
            print(f"ERRO: {e}")

    if not todos_dados:
        print("Nenhum dado extraído.")
        sys.exit(1)

    wb = Workbook()
    wb.remove(wb.active)

    criar_planilha_resumo(wb, todos_dados)
    criar_aba_formacao(wb, todos_dados)
    criar_aba_projetos(wb, "Proj. Pesquisa", todos_dados, "pesquisa")
    criar_aba_projetos(wb, "Proj. Ensino",   todos_dados, "ensino")
    criar_aba_projetos(wb, "Proj. Extensão", todos_dados, "extensao")

    for dados in todos_dados:
        criar_aba_pesquisador(wb, dados)

    saida = os.path.join(pasta, "curriculos_lattes.xlsx")
    wb.save(saida)
    print(f"\nPlanilha salva em: {os.path.abspath(saida)}")


if __name__ == "__main__":
    main()
